import psycopg2
from psycopg2.extras import RealDictCursor
from fuzzywuzzy import fuzz
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


class DataComparator:
    def __init__(self, db_config):
        self.db_config = db_config
        self.conn = None
        self.status_order = None

    def connect_to_db(self):
        self.conn = psycopg2.connect(**self.db_config, cursor_factory=RealDictCursor)
        self.status_order = self.get_status_order()

    def close_connection(self):
        if self.conn:
            self.conn.close()

    def get_status_order(self):
        with self.conn.cursor() as cur:
            cur.execute("SELECT status_name, status_value FROM statuses ORDER BY status_value")
            return {row['status_name']: row['status_value'] for row in cur.fetchall()}

    def check_staged_duplicates(self):
        with self.conn.cursor() as cur:
            cur.execute("""
                SELECT jcn, twcode, COUNT(*) as count
                FROM staged_egypt_weekly_data
                GROUP BY jcn, twcode
                HAVING COUNT(*) > 1
            """)
            return cur.fetchall()

    def compare_mrl_data(self):
        with self.conn.cursor() as cur:
            cur.execute("""
                SELECT 
                    s.staged_id,
                    s.jcn,
                    s.twcode,
                    s.nomenclature AS staged_nomenclature,
                    s.qty AS staged_qty,
                    s.ui AS staged_ui,
                    s.cog AS staged_cog,
                    s.fsc AS staged_fsc,
                    s.niin AS staged_niin,
                    s.part_no AS staged_part_no,
                    s.apl AS staged_apl,
                    m.order_line_item_id,
                    m.nomenclature AS mrl_nomenclature,
                    m.qty AS mrl_qty,
                    m.ui AS mrl_ui,
                    m.cog AS mrl_cog,
                    m.fsc AS mrl_fsc,
                    m.niin AS mrl_niin,
                    m.part_no AS mrl_part_no,
                    m.apl AS mrl_apl,
                    m.status_id AS mrl_status_id
                FROM staged_egypt_weekly_data s
                LEFT JOIN MRL_line_items m ON s.jcn = m.jcn AND s.twcode = m.twcode
            """)
            return cur.fetchall()

    def analyze_results(self, results, duplicates):
        summary = {
            "total_staged_records": len(results),
            "duplicate_jcn_twcode_in_staged": len(duplicates),
            "exact_jcn_twcode_matches": 0,
            "fuzzy_nomenclature_matches": 0,
            "exact_qty_matches": 0,
            "exact_ui_matches": 0,
            "exact_cog_matches": 0,
            "exact_fsc_matches": 0,
            "exact_niin_matches": 0,
            "exact_part_no_matches": 0,
            "exact_apl_matches": 0,
            "null_mismatches": 0,
            "data_mismatches": 0,
            "unmatched_records": 0,
        }

        categories = {
            "exact_match": [],
            "needs_update": [],
            "needs_manual_review": [],
            "unmatched": []
        }

        for row in results:
            result = {
                "staged_id": row["staged_id"],
                "jcn": row["jcn"],
                "twcode": row["twcode"],
                "order_line_item_id": row["order_line_item_id"],
            }

            if row["order_line_item_id"] is not None:
                summary["exact_jcn_twcode_matches"] += 1
                nomenclature_similarity = fuzz.ratio(row["staged_nomenclature"], row["mrl_nomenclature"])
                
                if nomenclature_similarity >= 80:
                    summary["fuzzy_nomenclature_matches"] += 1

                exact_matches = sum(row[f"staged_{field}"] == row[f"mrl_{field}"] for field in ["qty", "ui", "cog", "fsc", "niin", "part_no", "apl"])
                for field in ["qty", "ui", "cog", "fsc", "niin", "part_no", "apl"]:
                    if row[f"staged_{field}"] == row[f"mrl_{field}"]:
                        summary[f"exact_{field}_matches"] += 1
                    elif row[f"staged_{field}"] is None or row[f"mrl_{field}"] is None:
                        summary["null_mismatches"] += 1
                    else:
                        summary["data_mismatches"] += 1

                if nomenclature_similarity >= 80 and exact_matches == 7:
                    categories["exact_match"].append(result)
                elif nomenclature_similarity >= 80:
                    categories["needs_update"].append(result)
                else:
                    categories["needs_manual_review"].append(result)
            else:
                summary["unmatched_records"] += 1
                categories["unmatched"].append(result)

        return summary, categories

    def update_processing_categories(self, categories):
        with self.conn.cursor() as cur:
            for category, records in categories.items():
                staged_ids = [r["staged_id"] for r in records]
                cur.execute("""
                    UPDATE staged_egypt_weekly_data
                    SET processing_category = %s
                    WHERE staged_id = ANY(%s)
                """, (category, staged_ids))
        self.conn.commit()

    def process_records(comparator, results):
        for result in results:
            print(f"Processing record: JCN={result['jcn']}, TWCODE={result['twcode']}")
            print(f"Staged nomenclature: {result['staged_nomenclature']}")
            print(f"MRL nomenclature: {result['mrl_nomenclature']}")
            action = input("Enter 'u' to update, 'm' for manual processing, 's' to skip, or 'q' to quit: ").lower()
            
            if action == 'q':
                break
            elif action == 'u':
                # Implement update logic here
                print("Updating record...")
            elif action == 'm':
                print("Flagging for manual processing...")
            elif action == 's':
                print("Skipping record...")
            
            # Mark as processed
            with comparator.conn.cursor() as cur:
                cur.execute("""
                    UPDATE staged_egypt_weekly_data
                    SET processing_completed = TRUE
                    WHERE staged_id = %s
                """, (result['staged_id'],))
            comparator.conn.commit()

    def generate_excel_report(self, summary, categories, duplicates):
        wb = Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.append(["Metric", "Value"])
        for key, value in summary.items():
            ws_summary.append([key.replace("_", " ").title(), value])

        # Categories Sheet
        ws_categories = wb.create_sheet("Categories")
        ws_categories.append(["Category", "Staged ID", "JCN", "TWCODE", "MRL ID"])
        for category, records in categories.items():
            for record in records:
                ws_categories.append([category, record['staged_id'], record['jcn'], record['twcode'], record['order_line_item_id']])

        # Duplicates Sheet
        ws_duplicates = wb.create_sheet("Duplicates")
        ws_duplicates.append(["JCN", "TWCODE", "Count"])
        for dup in duplicates:
            ws_duplicates.append([dup['jcn'], dup['twcode'], dup['count']])

        # Apply some basic styling
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        return wb

def main():
    db_config = {
        "dbname": "ExtLogDB",
        "user": "postgres",
        "password": "123456",
        "host": "cmms-db-01",
        "port": "5432"
    }

    comparator = DataComparator(db_config)
    comparator.connect_to_db()

    try:
        # Check for duplicates in staged data
        duplicates = comparator.check_staged_duplicates()
        print(f"Found {len(duplicates)} duplicate JCN-TWCODE combinations in staged data.")

        results = comparator.compare_mrl_data()
        summary, categories = comparator.analyze_results(results, duplicates)
        
        # Update processing categories in the database
        comparator.update_processing_categories(categories)

        # Generate report
        wb = comparator.generate_excel_report(summary, categories, duplicates)
        wb.save("mrl_comparison_report.xlsx")
        print("\nFull report saved to mrl_comparison_report.xlsx")

        # Print summary of categorization
        for category, records in categories.items():
            print(f"{category}: {len(records)} records")

    finally:
        comparator.close_connection()

if __name__ == "__main__":
    main()